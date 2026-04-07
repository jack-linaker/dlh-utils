"""Functions for formatting to Pandas DataFrames.

Functions used to add visual formatting to Pandas dataframes and export them to Excel.
These wrap up some commonly-needed Pandas Styler and openpyxl boilerplate code.
"""

import os
import subprocess
from collections.abc import Callable
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from pandas.io.formats.style import Styler
from pyspark.sql import DataFrame


def apply_styles(
    df: pd.DataFrame | DataFrame, styles: dict[Callable[[Any], Any], str | list[str]]
) -> Styler:
    """Apply custom styles and return pandas `Styler`.

    Applies a set of custom style functions to a `DataFrame` and returns
    a pandas `Styler` object, which can be displayed by Jupyter and
    saved into Excel or HTML.

    Some suitable functions are included in this module with the prefix
    `style_` but you can also pass in a custom function.

    NOTE: This function returns a `Styler`, not a `DataFrame`.

    Parameters
    ----------
    df : `pandas.DataFrame` | `pyspark.sql.DataFrame`
        The `DataFrame` to be styled. If a Spark `DataFrame` is
        provided, it will be converted to Pandas.
    styles : `dict[Callable, str | list[str]]`
        A dictionary whose keys are functions and whose values are lists
        of column names. Each function should take in a single value and
        return a valid CSS string. The value can be a single column name
        (as a string) or a list

    Returns
    -------
    `pandas.io.formats.style.Styler`

    Examples
    --------
    The `DataFrame` `df` has a column, "Number", that can be positive or
    negative. We apply the default style to a column of this type using
    a `style_` function defined in this module:

    >>> apply_styles(df, {style_on_cutoff: "Number"})

    We would like to highlight in bold when a value is NA in two
    columns, "Number" and "OtherNumber". Both style rules will be
    applied to the "Number" column but only the bold style to
    "OtherNumber". Again, we use a `style_` function defined in this
    module:

    >>> apply_styles(
    ...     df,
    ...     {style_on_cutoff: "Number", style_on_condition: ["Number", "OtherNumber"]},
    ... )

    The `style_` functions have default behaviours we may want to
    customise. To do this, use the following pattern. The `partial`
    function is defined in `functools` and allows us to "freeze" some
    parameters of a function before it's evaluated:

    >>> from functools import partial
    >>> apply_styles(df, {partial(style_fill_pos_neg, property="color"): "Number"})
    """
    if not isinstance(df, pd.DataFrame):
        df = df.toPandas()

    # Build mapping: function -> list of columns.
    function_to_columns: dict[Callable[[Any], Any], list[str]] = {}
    for function, columns in styles.items():
        cols = [columns] if not isinstance(columns, list) else list(columns)

        # Validate each column exists.
        for c in cols:
            if c not in df.columns:
                error_message = f"apply_styles: column {c!r} not found in DataFrame."
                raise KeyError(error_message)

        function_to_columns.setdefault(function, []).extend(cols)

    sdf = df.style
    applied: list[tuple[Callable[[Any], Any], list[str]]] = []
    for function, columns in function_to_columns.items():
        unique_columns = list(dict.fromkeys(columns))

        # Apply style once for all columns for this function.
        sdf = sdf.map(function, subset=unique_columns)
        applied.append((function, unique_columns))

    # Attach metadata as attribute on the Styler.
    setattr(sdf, "_applied_styles", applied)

    return sdf


def copy_local_file_to_hdfs(
    local_path: str,
    hdfs_path: str,
    local_filename: str | None = None,
    hdfs_filename: str | None = None,
) -> None:
    """Copy a file created locally (ie in CDSW) to HDFS.

    Parameters
    ----------
    local_path : str
        Path to the local file to be copied.
    hdfs_path : str
        Target path to copy to.
    local_filename : str, optional
        If not specified, the local_path is assumed to include the
        filename. Defaults to None.
    hdfs_filename : str, optional
        If not specified, the hdfs_path is assumed to include the
        filename. Defaults to None.

    Examples
    --------
    >>> copy_local_file_to_hdfs("/tmp/wb.xlsx", "/hdfs/folder/xlsx")
    """
    if local_filename is not None:
        local_path = os.path.join(local_path, local_filename)
    if hdfs_filename is not None:
        hdfs_path = os.path.join(hdfs_path, hdfs_filename)
    commands = ["hadoop", "fs", "-put", "-f", local_path, hdfs_path]
    process = subprocess.Popen(
        commands,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    _stdout, _stderr = process.communicate()


def export_to_excel(
    dataframes: dict[str, DataFrame | pd.DataFrame]
    | list[DataFrame | pd.DataFrame]
    | DataFrame
    | pd.DataFrame,
    local_path: str,
    styles: dict[Any, Any] | None = None,
    columns: list[str] | dict[str, list[str]] | None = None,
    freeze_panes: dict[str, tuple[int, int]] | None = None,
    hdfs_path: str | None = None,
):
    """Create Excel workbook with one worksheet per provided DataFrame.

    Creates an Excel workbook with one worksheet for each of the
    provided DataFrames.

    Parameters
    ----------
    dataframes
        A dictionary whose keys are names for the sheets and values are
        Pandas or Spark DataFrames, or just a list of DataFrames; in the
        latter case the function will name the sheets Sheet1, Sheet2
        etc. If PySpark DataFrames are provided they will be converted
        to Pandas. A single DataFrame can also be passed for this
        argument.
    local_path : str
        The full path (including filename) where the Excel workbook will
        be saved. If not specified, the workbook will not be saved to
        disk.
    styles : optional
        A dictionary to pass to `apply_styles`. See the documentation
        for that function for more information. Defaults to None.
    columns : optional
        A dictionary whose keys are DataFrame names and whose values are
        lists of columns. If a DataFrame is named in this dictionary,
        only the listed columns will be written to Excel and their order
        will be as in the list provided. If a DataFrame is not named in
        this dictionary, all of its columns will be included in their
        default order. Defaults to None.
    freeze_panes : optional
        A dictionary mapping table names to tuples of the form (r, c)
        where r is the number of rows from the top to freeze and c the
        number of columns on the left. If a table's name is not present
        as a key, nothing will be frozen. Defaults to None.
    hdfs_path : str, optional
        The full HDFS path (including filename) where the Excel workbook
        will be saved. If you specify this, you must also provide a
        `local_path`. Defaults to None.

    Returns
    -------
    `openpyxl.WorkBook`

    Examples
    --------
    Write two DataFrames to named sheets, selecting only two columns
    from `people_df`:

    >>> export_to_excel(
    ...     {"People": people_df, "Places": places_df},
    ...     "/tmp/abc.xlsx",
    ...     columns={"People": ["surname", "firstname"]},
    ... )

    This next example shows a single DataFrame with complex styling:
        - bank_balance is given a background colour gradient.
        - income_change is given red/green text colours for
          negative/positive.
        - age is formatted in bold if the person is a child.

    Note the use of `partial` to set values for the parameters when we
    don't want to use their default values.

    See `apply_styles` and the `style_*` functions for more information
    on these functions.

    >>> from functools import partial
    >>> export_to_excel(
    ...     {"People": people_df},
    ...     "/tmp/abc.xlsx",
    ...     styles={
    ...         partial(
    ...             style_colour_gradient,
    ...             min=df["bank_balance"].min(),
    ...             max=df["bank_balance"].max(),
    ...         ): "bank_balance",
    ...         partial(style_on_cutoff, property="color"): "income_change",
    ...         partial(style_on_condition, condition=lambda x: x < 18): "age",
    ...     },
    ... )
    """
    # Normalise `dataframes` into an ordered mapping: sheet_name ->
    # pandas.DataFrame.
    if isinstance(dataframes, pd.DataFrame):
        dataframes = {"Sheet1": dataframes}
    elif isinstance(dataframes, list):
        # Name sheets Sheet1, Sheet2 etc.
        dataframes = {f"Sheet{i}": df for i, df in enumerate(dataframes, start=1)}
    elif not isinstance(dataframes, dict):
        # Last resort: try to treat as a single df-like object.
        dataframes = {"Sheet1": dataframes}

    # If `columns` is provided as a `list` but the user passed a single
    # sheet mapping, associate the list with that single sheet's name.
    if isinstance(columns, list):
        if len(dataframes) > 1:
            error_message = (
                "Can't pass a list of columns to write_excel unless you only passed in"
                " a single dataframe. You can use a dictionary instead (see this "
                "function's docstring for an example)."
            )
            raise ValueError(error_message)
        only_sheet = next(iter(dataframes.keys()))
        columns = {only_sheet: columns}

    columns = columns or {}
    styles = styles or {}
    freeze_panes = freeze_panes or {}

    # Ensure the directory exists.
    Path.mkdir(Path(local_path).parent, parents=True, exist_ok=True)

    # Use `ExcelWriter` as a context manager (ensures write is closed
    # and file saved).
    with pd.ExcelWriter(local_path, engine="openpyxl", mode="w") as writer:
        # Iterate sheets in the user-specified order.
        for sheet_name, df_like in dataframes.items():
            # Convert Spark DataFrames.
            if not isinstance(df_like, pd.DataFrame):
                if hasattr(df_like, "toPandas"):
                    df = df_like.toPandas()
                else:
                    df = pd.DataFrame(df_like)
            else:
                df = df_like

            # Subset columns if requested for this sheet.
            if sheet_name in columns:
                df = df.loc[:, columns[sheet_name]]

            # Default: nothing to write yet.
            df_to_write = None

            # Apply styles if provided for this sheet.
            if sheet_name in styles:
                # Expect `styles[sheet_name]` to be a `dict` for
                # `apply_styles`.
                style_or_df = apply_styles(df, styles[sheet_name])
                # `apply_styles` is expected to return `pandas.Styler`.
                if not isinstance(style_or_df, Styler):
                    # Fallback: treat as DataFrame.
                    df_to_write = style_or_df
                else:
                    # Write Styler to excel.
                    style_or_df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                df_to_write = df

            if df_to_write is not None:
                df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)

            # Ensure the worksheet object is available for tweaks
            # (freeze_panes etc).
            ws = writer.book[sheet_name]

            # Apply freeze panes if requested.
            if sheet_name in freeze_panes:
                r, c = freeze_panes[sheet_name]
                # `openpyxl` expects the first unfrozen cell; we set via
                # a cell object and freeze r rows and c columns => first
                # unfrozen is (r+1, c+1).
                ws.freeze_panes = ws.cell(row=r + 1, column=c + 1)

        # Remove default sheet if present and not requested.
        if "Sheet" in writer.book.sheetnames and "Sheet" not in dataframes:
            del writer.book["Sheet"]

    # Optionally copy to HDFS.
    if hdfs_path is not None:
        copy_local_file_to_hdfs(local_path, hdfs_path)

    return openpyxl.load_workbook(local_path)


def style_colour_gradient(
    value: Any,
    min: Any,
    max: Any,
    property: str = "background-color",
    min_colour: str = "#FFFFFF",
    max_colour: str = "#FF0000",
    error_colour: str | None = "#000000",
) -> str:
    """Map numeric value to gradient colour CSS string.

    Returns a CSS string that sets the specified colour property to a
    colour ranging between start_colour and end_colour depending on the
    value's position in the range between min and max.

    This function is intended to be used with apply_styles() (defined in
    this module).

    Parameters
    ----------
    value : any numeric type
        The value to be mapped to a colour.
    min : any numeric type
        The highest value that will receive the start_colour (any lower
        values also receive start_colour).
    max : any numeric type
        The lowest value that will receive the end_colour (any higher
        values also receive end_colour).
    property : str, optional
        The CSS property the style will be applied to. Must be able to
        be set to a hexadecimal colour string. Defaults to
        "background-color".
    min_colour : str, optional
        The colour at the minimum end of the gradient. Pass only a
        hexadecimal string, not a colour name. Defaults to "#FFFFFF".
    max_colour : str, optional
        The colour at the maximum end of the gradient. Pass only a
        hexadecimal string, not a colour name. Defaults to "#FF0000".
    error_colour : str, optional
        The colour will be assigned if an error occurs in this function.
        If None, the error will be raised instead. Defaults to
        "#000000".

    Returns
    -------
    str
    """
    try:
        # Extract colour channels from parameters
        min_colour = min_colour.replace("#", "")
        max_colour = max_colour.replace("#", "")
        min_channels = [int(min_colour[i : i + 2], 16) for i in (0, 2, 4)]
        max_channels = [int(max_colour[i : i + 2], 16) for i in (0, 2, 4)]

        # Interpolate
        position = (value - min) / (max - min)
        interpolated_channels = [0, 0, 0]
        for c in range(3):
            if max_channels[c] > min_channels[c]:
                val = int(
                    position * (max_channels[c] - min_channels[c]) + min_channels[c]
                )
            else:
                val = int(
                    (1 - position) * (min_channels[c] - max_channels[c])
                    + max_channels[c]
                )
            interpolated_channels[c] = ("0x%0*x" % (2, val))[2:].upper()

        # Return the result
        return property + " : #" + "".join(interpolated_channels) + ";"

    except Exception as ex:
        if error_colour is None:
            raise ex
        return property + " : #" + error_colour + ";"


def style_map_values(
    value: Any,
    mapping_dictionary: dict[Any, Any],
    property: str = "background-color",
    default_style: str | None = None,
    error_style: str | None = None,
) -> str:
    """Map values to CSS styles using a lookup dictionary.

    Returns a CSS string that sets the specified property to a value as
    specified by mapping_dictionary, which maps possible values being
    passed in to the style the property should be assigned to. If the
    value is not found in mapping_dictionary the default_value is used,
    if one is specified, or an error is raised. In the event of an
    error, error_value will be used if it is not None, otherwise the
    caller will receive the error.

    This function is intended to be used with apply_styles() (defined in
    this module).

    Parameters
    ----------
    value : any appropriate type
        A value of a type that can be accepted by the condition
        function.
    mapping_dictionary : dict
        Keys are possible values for the parameter "value"; these are
        mapped to styles.
    property : str, optional
        The CSS property the style will be applied to. Must be able to
        be set to a hexadecimal colour string. Defaults to
        "background-color".
    default_style : str, optional
        If not None, this will be used if the value passed in is not
        found in mapping_dictionary. Defaults to None.
    error_style : str, optional
        The style will be assigned if an error occurs in this function.
        If None, the error will be raised instead. Defaults to None.

    Returns
    -------
    str
    """
    try:
        if value in mapping_dictionary:
            style = mapping_dictionary[value]
        elif default_style is not None:
            style = default_style
        else:
            str_val = str(value)
            error_message = (
                f"Value {str_val} not found in mapping_dictionary and no default_value"
                " was specified."
            )
            raise ValueError(error_message)
        # Return the result
        return property + " : " + style + ";"

    except Exception as ex:
        if error_style is None:
            raise ex
        return property + " : " + error_style + ";"


def style_on_condition(
    value: Any,
    property: str = "font-weight",
    true_style: str = "bold",
    false_style: str = "normal",
    error_style: str | None = None,
    condition: Any = lambda x: x == 0,
) -> str:
    """Return CSS style string based on a condition applied to a value.

    Returns a CSS string that sets the specified property to the
    appropriate style for the value passed in.

    This function is intended to be used with apply_styles() (defined in
    this module).

    Parameters
    ----------
    value : any appropriate type
        A value of a type that can be accepted by the condition
        function.
    property : str, optional
        The CSS property the style will be applied to. Defaults to
        "font-weight".
    true_style : str, optional
        The style will be assigned when the condition evaluates true on
        the value. Defaults to "bold".
    false_style : str, optional
        The style will be assigned when the condition evaluates false on
        the value. Defaults to "normal".
    error_style : str, optional
        The style will be assigned if an error occurs in this function.
        If None, the error will be raised instead. Defaults to None.
    condition : function, optional
        A function that accepts value and returns a truthy value. This
        is used to determine whether the current value receives
        true_style or false_style. The default function applies
        true_style to values that exactly equal zero. Defaults to lambda
        x: x == 0.

    Returns
    -------
    str
    """
    try:
        if condition(value):
            return property + " : " + true_style + ";"
        return property + " : " + false_style + ";"
    except Exception as ex:
        if error_style is None:
            raise ex
        return property + " : " + error_style + ";"


def style_on_cutoff(
    value: Any,
    cutoff: Any = 0,
    negative_style: str = "red",
    positive_style: str = "green",
    zero_style: str = "white",
    error_style: str | None = "black",
    property: str = "background-color",
) -> str:
    """Return CSS style string based on value compared to cutoff.

    Returns a CSS string that sets the specified property to the
    appropriate style for the value passed in. The style is chosen based
    on whether the value is greater than, equal to or less than the
    cutoff.

    By default the cutoff is 0 and the function assigns background
    colours: green for positive, red for negative and white for exactly
    zero.

    You can also set a style for if the attempt to calculate a result
    led to an exception or if value < cutoff, value > cutoff and value
    == cutoff all evaluate to False.

    This function is intended to be used with apply_styles() (defined in
    this module).

    Parameters
    ----------
    value : numeric or other appropriate type
        A value of any type that can be compared with cutoff using "<"
        and ">".
    cutoff : any type comparable to value, optional
        The passed-in value will be compared to cutoff to determine
        which style is returned. Defaults to 0.
    negative_style : str, optional
        The colour name, RGB code or other style value to be assigned
        when value < cutoff. Defaults to "red".
    positive_style : str, optional
        The colour name, RGB code or other style value to be assigned
        when value > cutoff. Defaults to "green".
    zero_style : str, optional
        The colour name, RGB code or other style value to be assigned
        when neither value < cutoff nor value > cutoff. Defaults to
        "white".
    error_style : str, optional
       The colour name, RGB code or other style value to be assigned
       when an error occurs. This can happen when the value is NaN or
       not of the expected type. If error_style=None, the exception will
       be re-raised instead. If in doubt, pass in None and make sure any
       errors that are raised are expected. Defaults to "black".
    property : str, optional
        The CSS property the colour will be applied to. Defaults to
        "background-color".

    Returns
    -------
    str
    """
    try:
        if value < cutoff:
            return property + " : " + negative_style + ";"
        elif value > cutoff:
            return property + " : " + positive_style + ";"
        elif value == cutoff:
            return property + " : " + zero_style + ";"
        else:
            error_message = (
                f"Value {value} was not less than, equal to, or greater than cutoff "
                f"{cutoff}."
            )
            raise ValueError(error_message)
    except Exception as ex:
        if error_style is None:
            raise ex
        return property + " : " + error_style + ";"
